#!/usr/bin/env python3
'''
Geoffrey Weal, Adsorber_Part_D_create_excel_file.py, 22/08/2021

Create excel file that contains all the information from your jobs

'''
import os, sys, string

from ase import Atoms
from ase.io import read, write

from Adsorber import __version__
from Adsorber.Subsidiary_Programs.Part_D_Methods import introductory_remarks, get_project_id_and_time_from_slurm, get_start_date_from_OUTCAR, determine_convergence_and_time_elapsed_and_date_finished_and_Max_mem_Gb_and_energy_of_output
from Adsorber.Subsidiary_Programs.Part_D_Methods import get_job_id, get_EDIFFG_from_OUTCAR
from Adsorber.Subsidiary_Programs.Part_D_Methods import get_cluster_name_from_Run_AdsorberPY_script

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#from openpyxl.styles import colors
from openpyxl.styles import Alignment # , Color, Font
center_alignment = Alignment(horizontal='center')
from openpyxl.styles.fills import PatternFill

if len(sys.argv) <= 1:
	get_adsorbates = 'all'
else:
	get_adsorbates = sys.argv[1:]

introductory_remarks()

System_Adsorbate_VASP_Jobs_folder_name = 'Part_C_Selected_Systems_with_Adsorbed_Species_to_Run_in_VASP'
if not System_Adsorbate_VASP_Jobs_folder_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(System_Adsorbate_VASP_Jobs_folder_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+System_Adsorbate_VASP_Jobs_folder_name+' folder is.')
    print('This program will finish without beginning.')
    exit()

Run_AdsorberPY_name = 'Run_Adsorber.py'
if not Run_AdsorberPY_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(Run_AdsorberPY_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+Run_AdsorberPY_name+' script is.')
    print('This program will finish without beginning.')
    exit()

name_of_xyz_file = get_cluster_name_from_Run_AdsorberPY_script(Run_AdsorberPY_name)
bare_system_xyz = read(name_of_xyz_file)
no_of_atoms_in_bare_system_xyz = len(bare_system_xyz)

# --------------------------------------------------------------------------------------------------------------------------

Submission_Folder = 'Submission_Folder'
def add_to_data(data, root, folder_name):
	root_split = root.split('/')
	project_id, time_submitted_for = get_project_id_and_time_from_slurm(root)
	project_name = root_split[-1]
	if folder_name == 'Part_A_Non_Adsorbed_Files_For_VASP':
		sheet_name = 'Originals'
	else:
		sheet_name = project_name.split('_')[0]
	description = root.replace(folder_name+'/','')
	print(description)
	submission_folders = [dirname for dirname in os.listdir(root) if os.path.isdir(root+'/'+dirname) and dirname.startswith(Submission_Folder)]
	path_to_OUTCAR = None; files_in_submission_folder = None
	if os.path.exists(root+'/OUTCAR'):
		path_to_OUTCAR = root
	elif len(submission_folders) > 0:
		# --------
		# get the name of the highest submission folder
		submission_folders_values = []
		for submission_folder_name in submission_folders:
			if os.path.exists(root+'/'+submission_folder_name+'/OUTCAR'):
				if 'Issue' in submission_folder_name:
					numbers = tuple(int(value) for value in submission_folder_name.replace(Submission_Folder+'_','').split('_Issue_'))
				else:
					numbers = (int(submission_folder_name.replace(Submission_Folder+'_','')),0)
				submission_folders_values.append(numbers)
		highest_submission_folder_value = max(submission_folders_values)
		if highest_submission_folder_value[1] == 0:
			highest_submission_folder = Submission_Folder+'_'+str(highest_submission_folder_value[0])
		else:
			highest_submission_folder = Submission_Folder+'_'+str(highest_submission_folder_value[0])+'_Issue_'+str(highest_submission_folder_value[1])
		# --------
		if os.path.exists(root+'/'+highest_submission_folder+'/OUTCAR'):
			files_in_submission_folder = highest_submission_folder
			path_to_OUTCAR = root+'/'+highest_submission_folder
	found_OUTCAR = path_to_OUTCAR is not None
	if found_OUTCAR:
		job_id = get_job_id(path_to_OUTCAR)
		date_submitted, start_time_timestamp = get_start_date_from_OUTCAR(path_to_OUTCAR)
		did_converge, time_elapsed, date_finished, Max_mem_Gb, energy = determine_convergence_and_time_elapsed_and_date_finished_and_Max_mem_Gb_and_energy_of_output(path_to_OUTCAR,start_time_timestamp)
		EDIFFG = get_EDIFFG_from_OUTCAR(path_to_OUTCAR)
	else:
		job_id = 'None'
		did_converge = 'OUTCAR not found'
		date_submitted, start_time_timestamp, time_elapsed, date_finished, Max_mem_Gb, energy = None, None, None, None, None, None
		EDIFFG = '---'
	notes = ''
	data.setdefault(sheet_name,[]).append([job_id,project_id,project_name,root,files_in_submission_folder,description,time_submitted_for,date_submitted,time_elapsed,date_finished,Max_mem_Gb,EDIFFG,energy,did_converge,None,None,notes])
	return found_OUTCAR, description

def check_subdirectories(data, path, folder_name, get_adsorbates, OUTCAR_not_found, jobs_not_run):
	submitSL_file = 'submit.sl'
	INCAR_file = 'INCAR'
	submission_folder_name = 'Submission_Folder'
	OUTCAR_file = 'OUTCAR'
	for root, dirs, files in os.walk(path):
		if submission_folder_name in root:
			dirs[:] = []
			files[:] = []
			continue
		for index in range(len(dirs)-1,-1,-1):
			dirname = dirs[index]
			if dirname.startswith(submission_folder_name):
				del dirs[index]
		if (submitSL_file in files) and (INCAR_file in files):
			found_OUTCAR, description = add_to_data(data, root, folder_name)
			if not (OUTCAR_file in files):
				if (folder_name == 'Part_A_Non_Adsorbed_Files_For_VASP') and not ('system' in root):
					pass
				else:
					OUTCAR_not_found.append(description)
			files[:] = []
			dirs[:] = []
			continue

folder_names = ['Part_A_Non_Adsorbed_Files_For_VASP','Part_C_Selected_Systems_with_Adsorbed_Species_to_Run_in_VASP']
print('==================================================')
print('Collecting data about OUTCAR files in: '+str(folder_names))
data = {}
OUTCAR_not_found = []
jobs_not_run = []
for folder_name in folder_names:
	directories = [folder for folder in os.listdir(folder_name) if os.path.isdir(folder_name+'/'+folder)]
	if (not get_adsorbates == 'all') and (folder_name == 'Part_C_Selected_Systems_with_Adsorbed_Species_to_Run_in_VASP'):
		for index in range(len(directories)-1,-1,-1):
			folder = directories[index]
			if not any([(folder.split('_')[0] == adsorbate_name) for adsorbate_name in get_adsorbates]):
				del directories[index]
	directories.sort()
	for directory in directories:
		path = folder_name+'/'+directory
		check_subdirectories(data, path, folder_name, get_adsorbates, OUTCAR_not_found, jobs_not_run)


if not len(OUTCAR_not_found) == 0:
	print('-----------------------------------------------------------------')
	print('Note: the following do not have an OUTCAR file and therefore may have not been run on VASP+slurm.')
	print('This may be because they have just starting running on slurm if these jobs are currently running on slurm')
	for path in OUTCAR_not_found:
		print(path)
	print('(scroll up for start of message)')
	input_string = input('Press enter to continue, or x then enter to close this program: ')
	if input_string in ['x']:
		exit('Closing this program')
	print('-----------------------------------------------------------------')

# --------------------------------------------------------------------------------------------------------------------------

print('==================================================')
print('Performing structural analysis.')
from datetime import timedelta
from tqdm import tqdm, trange
from Adsorber.Subsidiary_Programs.Part_D_Methods import get_OUTCAR_Atoms_files, compare_position_of_adsorbates, make_graph_of_adsorbate, rotation_second_system

def comprehensive_structural_analysis(data,sheet_name,all_outcar_objects,no_of_atoms_in_bare_system_xyz,attach_edges=True,compare_neighbours=True):
	print('-----------------------------------------------------')
	print('Making graphs of system+adsorbate objects '+str(sheet_name))
	graphs_of_system_adsorbates = {}
	for system_adsorbate, name, path in tqdm(all_outcar_objects,unit="graph"):
		graph = make_graph_of_adsorbate(system_adsorbate,no_of_atoms_in_bare_system_xyz,attach_edges=attach_edges)
		graphs_of_system_adsorbates[name] = graph
	# determine similar systems
	print('-----------------------------------------------------')
	print('Comparing structural similarity between system+adsorbates for '+str(len(all_outcar_objects))+' data entries, ')
	print('and also determining which atoms on the adsorbate are bound to which atoms on your system')
	print('Note that this process may be slow in the beginning, but gets faster as it proceeds.')
	similar_system_adsorbates = {}
	counter = -1
	for index1 in trange(len(all_outcar_objects),unit="systems processed"):
		first_system_adsorbate, first_name, first_path = all_outcar_objects[index1]
		first_graph = graphs_of_system_adsorbates[first_name]
		masses = first_system_adsorbate.get_masses()
		for index2 in range(index1+1,len(all_outcar_objects)):
			second_system_adsorbate, second_name, second_path = all_outcar_objects[index2]
			second_graph = graphs_of_system_adsorbates[second_name]
			second_system_adsorbate_copy = rotation_second_system(first_system_adsorbate,second_system_adsorbate,no_of_atoms_in_bare_system_xyz,masses)
			if compare_position_of_adsorbates(first_system_adsorbate,first_graph,second_system_adsorbate_copy,second_graph,no_of_atoms_in_bare_system_xyz,compare_neighbours=compare_neighbours):
				similar_system_adsorbates.setdefault(first_name,[]).append(second_name)
				similar_system_adsorbates.setdefault(second_name,[]).append(first_name)
		while counter < len(data[sheet_name]):
			counter += 1
			if data[sheet_name][counter][2] == first_name:
				break
		data[sheet_name][counter][-3] = similar_system_adsorbates.get(first_name,None)
		neighbours = {}
		for index in first_graph.nodes:
			datum = first_graph.nodes[index]
			neighbours.setdefault(datum['symbol'],[]).append((index,(datum['system_neighbours'])))
		data[sheet_name][counter][-2] = neighbours

from Adsorber.Subsidiary_Programs.Part_D_Methods import nnd_and_half_distance
from Adsorber.Adsorber.Part_B_adsorb_single_species_to_cluster import is_same_system, get_distance
def get_neighbours(system_adsorbate,len_of_cluster):
	neighbours = {}
	sys_atom_positions = system_adsorbate.get_positions()[:len_of_cluster]
	for abs_index in range(len_of_cluster,len(system_adsorbate)):
		adsorbate_atom = system_adsorbate[abs_index]
		adsorbate_symbol = adsorbate_atom.symbol
		system_neighbours = []
		for sys_index in range(len_of_cluster):
			distance = get_distance(adsorbate_atom.position,sys_atom_positions[sys_index])
			if distance <= nnd_and_half_distance(adsorbate_symbol,system_adsorbate[sys_index].symbol):
				system_neighbours.append(sys_index)
		neighbours.setdefault(adsorbate_symbol,[]).append((abs_index,system_neighbours))
	return neighbours

def simple_adsorbate_structural_analysis(data,sheet_name,all_outcar_objects,len_of_system):
	print('-----------------------------------------------------')
	print('Determining if adsorbates between systems are in proximity to each other')
	print('and also determining which atoms on the adsorbate are bound to which atoms on your system')
	counter = -1
	similar_system_adsorbates = {}
	for index1 in trange(len(all_outcar_objects),unit="systems processed"):
		first_system_adsorbate, first_name, first_path = all_outcar_objects[index1]
		masses = first_system_adsorbate.get_masses()
		for index2 in range(index1+1,len(all_outcar_objects)):
			second_system_adsorbate, second_name, second_path = all_outcar_objects[index2]
			second_system_adsorbate_copy = rotation_second_system(first_system_adsorbate,second_system_adsorbate,len_of_system,masses)
			if is_same_system(first_system_adsorbate,second_system_adsorbate_copy,starting_index=len_of_system,maximum_distance_to_be_within=1.0):
				similar_system_adsorbates.setdefault(first_name,[]).append(second_name)
				similar_system_adsorbates.setdefault(second_name,[]).append(first_name)
		while counter < len(data[sheet_name]):
			counter += 1
			if data[sheet_name][counter][2] == first_name:
				break
		data[sheet_name][index1][-3] = similar_system_adsorbates.get(first_name,None)
		# get neighbouring atoms
		data[sheet_name][index1][-2] = get_neighbours(first_system_adsorbate,len_of_system)

def no_structural_analysis(data,sheet_name,all_outcar_objects,len_of_system):
	print('-----------------------------------------------------')
	print('Determining which atoms on the adsorbate are bound to which atoms on your system')
	counter = -1
	for index1 in trange(len(all_outcar_objects),unit="systems processed"):
		first_system_adsorbate, first_name, first_path = all_outcar_objects[index1]
		while counter < len(data[sheet_name]):
			counter += 1
			if data[sheet_name][counter][2] == first_name:
				break
		neighbours = get_neighbours(first_system_adsorbate,len_of_system)
		if len(neighbours) == 0:
			print('Warning: The adsorbate may not be attached to the surface of your system in '+str(first_name)+' ('+str(first_path)+').')
		data[sheet_name][index1][-2] = neighbours

def convert_neighbours_to_string(neighbours):
	if neighbours is None:
		return '---'
	details = []
	for symbol, atoms in neighbours.items():
		counter = 1
		for atom in atoms:
			detail = [atom[1],symbol,counter,atom[0]]
			details.append(detail)
	details.sort(key=lambda x:(-len(x[0]),x[1],x[2]))
	a_string = [str(len(neighs))+' ('+symbol+str(a_no)+' ['+str(index)+'->'+','.join([str(nn) for nn in neighs])+'])' for neighs,symbol,a_no,index in details if len(neighs) > 0]
	a_string = ','.join(a_string)
	return a_string

def look_for_similar_systems(data,sheet_name):
	similar_system_adsorbates = {}
	for datum in data[sheet_name]:
		neighbours = convert_neighbours_to_string(datum[-2])
		similar_system_adsorbates[neighbours] = similar_system_adsorbates.get(neighbours,0) + 1
	for index in range(len(data[sheet_name])):
		neighbours = convert_neighbours_to_string(data[sheet_name][index][-2])
		data[sheet_name][index][-3] = similar_system_adsorbates[neighbours]

for sheet_name, data_for_sheet in data.items():
	if sheet_name == 'Originals':
		continue
	print('-----------------------------------------------------')
	print('Getting system+adsorbate objects by processing CONTCARs/OUTCARs from '+str(sheet_name))
	all_outcar_objects = []
	for index in trange(len(data_for_sheet),unit="CONTCAR or OUTCAR"):
		job_id,project_id,project_name,path_to_VASP_folder,files_in_submission_folder,description,time_submitted_for,date_submitted,time_elapsed,date_finished,maximum_memory_used,EDIFFG,energy,did_converge,similar_systems,neighbours,notes = data_for_sheet[index]
		outcar_location_data = get_OUTCAR_Atoms_files(path_to_VASP_folder,files_in_submission_folder,no_of_atoms_in_bare_system_xyz)
		if outcar_location_data is None:
			continue
		all_outcar_objects.append(outcar_location_data)
	# make graphs of plots
	no_of_entries = len(all_outcar_objects)
	'''
	if no_of_entries <= 500:
		attach_edges = True; compare_neighbours = True
		comprehensive_structural_analysis(data,sheet_name,all_outcar_objects,no_of_atoms_in_bare_system_xyz,attach_edges=attach_edges,compare_neighbours=compare_neighbours)
	elif no_of_entries <= 2000:
		simple_adsorbate_structural_analysis(data,sheet_name,all_outcar_objects,no_of_atoms_in_bare_system_xyz)
	else:
		no_structural_analysis(data,sheet_name,all_outcar_objects,no_of_atoms_in_bare_system_xyz)
	'''
	no_structural_analysis(data,sheet_name,all_outcar_objects,no_of_atoms_in_bare_system_xyz)
	look_for_similar_systems(data,sheet_name)
	print('-----------------------------------------------------')

# --------------------------------------------------------------------------------------------------------------------------

print('==================================================')
print('Placing data into excel spreadsheet.')

green_colour = PatternFill("solid", fgColor="0099CC00")
red_colour   = PatternFill("solid", fgColor="00FF6600")

def convert_similar_systems_to_string(similar_systems):
	if similar_systems is None:
		return '---'
	if isinstance(similar_systems,int):
		return similar_systems
	elif isinstance(similar_systems,list):
		return ', '.join(similar_systems)
	else:
		print('check')
		import pdb; pdb.set_trace()

def get_key_name(key):
	key_split = key.replace(' ','_')
	key_split = key_split.rstrip().split('(,')
	for index in range(len(key_split)):
		key_split_part = key_split[index].replace('[','')
		key_split_part = key_split_part.replace(']','')
		key_split_part = key_split_part.replace('(','')
		key_split_part = key_split_part.replace(')','')
		key_split_part = key_split_part.replace('->','to')
		key_split[index] = key_split_part.replace(',','+')
	key_name = '__'.join(key_split)
	return key_name

workbook = Workbook()
for sheet_name, data_for_sheet in data.items():
	print(sheet_name)
	sheet = workbook.create_sheet(sheet_name)
	title = ['Job', 'Project', 'Job Name', 'Path', 'Description', 'Time submitted for', 'Date Submitted', 'Date Finished', 'Time Elapsed (hrs)', 'Max. Memory (Gb)', 'EDIFFG (eV)', 'Energy (eV)', 'Rel. Energy (eV)', 'Converged', 'Similar to', 'No of surface atoms adsorbed to', 'Folder name in "Similar_Systems"', 'Notes']
	for index in range(len(title)):
		sheet[get_column_letter(index+1)+'1'] = title[index]
		sheet[get_column_letter(index+1)+'1'].alignment = center_alignment
	#data_for_sheet.sort(key=lambda datum: datum[3])
	#import pdb; pdb.set_trace()
	data_for_sheet.sort(key=lambda datum: float('inf') if datum[12] is None else datum[12])
	#energies = [(datum[12], index_in_excel, datum[3]) for (datum, index_in_excel) in zip(data_for_sheet, range(2,len(data_for_sheet)+2)) if datum[12] is not None]
	#min_energy, min_energy_index_in_excel, _ = min(energies,key=lambda datum: datum[0])
	for index in range(len(data_for_sheet)):
		job_id,project_id,project_name,path_to_VASP_folder,files_in_submission_folder,description,time_submitted_for,date_submitted,time_elapsed,date_finished,maximum_memory_used,EDIFFG,energy,did_converge,similar_systems,neighbours,notes = data_for_sheet[index]
		sheet['A'+str(index+2)] = job_id
		sheet['B'+str(index+2)] = project_id
		sheet['C'+str(index+2)] = project_name
		sheet['D'+str(index+2)] = path_to_VASP_folder
		sheet['E'+str(index+2)] = description
		sheet['F'+str(index+2)] = time_submitted_for
		sheet['G'+str(index+2)] = date_submitted
		sheet['H'+str(index+2)] = date_finished
		sheet['I'+str(index+2)] = time_elapsed
		sheet['J'+str(index+2)] = '='+str(maximum_memory_used) if (maximum_memory_used is not None) else None
		sheet['K'+str(index+2)] = EDIFFG
		if not energy is None:
			energy = round(energy,14)
			sheet['L'+str(index+2)] = '='+str(energy)
			if sheet_name == 'Originals':
				sheet['M'+str(index+2)] = ''
			else:
				sheet['M'+str(index+2)] = '=L'+str(index+2)+'-MIN(L:L)'
		else:
			sheet['L'+str(index+2)] = sheet['L'+str(index+2)] = ''
		sheet['N'+str(index+2)] ='Yes' if did_converge else 'No'
		sheet['N'+str(index+2)].fill = green_colour if did_converge else red_colour
		sheet['O'+str(index+2)] = convert_similar_systems_to_string(similar_systems)
		sheet['P'+str(index+2)] = convert_neighbours_to_string(neighbours)
		sheet['Q'+str(index+2)] = get_key_name(sheet['P'+str(index+2)].value)
		sheet['R'+str(index+2)] = notes
		for letter in string.ascii_uppercase[:13]:
			sheet[letter+str(index+2)].alignment = center_alignment

std = workbook.get_sheet_by_name('Sheet')
workbook.remove_sheet(std)

save_folder_name = 'Part_D_Results_Folder'
def make_dir(save_folder_name):
	if not os.path.exists(save_folder_name):
		os.makedirs(save_folder_name)
make_dir(save_folder_name)

excel_name = "Part_D_Information_on_VASP_Calculations.xlsx"
print('Saving excel spreadsheet as: '+str(excel_name))
workbook.save(filename=save_folder_name+'/'+excel_name)
print('Spreadsheet has been saved.')

# --------------------------------------------------------------------------------------------------------------------------

def get_OUTCAR_file(path_to,index):
	try:
		OUTCAR_images = read(path_to,index=index)
	except:
		OUTCAR_images = read(path_to+'.gz',index=index)
	return OUTCAR_images

print('==================================================')
print('Placing data into making traj files for each site.')
for sheet_name, data_for_sheet in data.items():
	print('--------------------------------------------------')
	print(sheet_name)
	print('--------------------------------------------------')
	energies = [(datum[11], index_in_excel, datum[3]) for (datum, index_in_excel) in zip(data_for_sheet, range(2,len(data_for_sheet)+2)) if datum[10] is not None]
	# Make traj files of systems
	energies.sort(key=lambda datum: datum[0])
	if sheet_name == 'Originals':
		folder_name = folder_names[0]
	else:
		folder_name = folder_names[1]
	path_to_place_traj = save_folder_name+'/VASP_job_trajectories/'+sheet_name
	make_dir(path_to_place_traj)
	counter = 1
	pbar = tqdm(energies,unit="traj files processed")
	for energy, index_in_excel, path_to_VASP_folder in pbar:
		pbar.set_description("Processing %s" % str(path_to_VASP_folder.split('/')[-1]))
		folders = [folder for folder in os.listdir(path_to_VASP_folder) if (os.path.isdir(path_to_VASP_folder+'/'+folder) and folder.startswith(Submission_Folder) and not ('Issue' in folder))]
		folders.sort(key=lambda x:int(x.replace(Submission_Folder+'_','')))
		minimisation_process = []
		for index in range(len(folders)):
			folder = folders[index]
			if index == 0:
				minimisation_process += get_OUTCAR_file(path_to_VASP_folder+'/'+folder+'/OUTCAR',index=':')
			else:
				minimisation_process += get_OUTCAR_file(path_to_VASP_folder+'/'+folder+'/OUTCAR',index='1:')
		if len(folders) == 0:
			minimisation_process += get_OUTCAR_file(path_to_VASP_folder+'/OUTCAR',index=':')
		else:
			minimisation_process += get_OUTCAR_file(path_to_VASP_folder+'/OUTCAR',index='1:')
		if not (0 <= index_in_excel-2 <= len(data_for_sheet)):
			print('Error')
			import pdb; pdb.set_trace()
			exit()
		job_name = data_for_sheet[index_in_excel-2][2]
		if sheet_name == 'Originals':
			path_and_name = path_to_place_traj+'/'+str(job_name)+'_FT.traj'
		else:
			path_and_name = path_to_place_traj+'/'+str(counter)+'_'+str(job_name)+'_FT.traj'
		write(path_and_name, minimisation_process)
		counter += 1

print('All traj files have been written.')
print('Done')
print('==================================================')
