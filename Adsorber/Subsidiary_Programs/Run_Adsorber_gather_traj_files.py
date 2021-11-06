#!/usr/bin/env python3
'''
Geoffrey Weal, Run_Adsorber_compare_sytems_with_same_binding.py, 14/09/2021

This program is designed to allow you to look at different VASP jobs with adsorbates in different positions that have converged to similar places.

This program is designed to allow you to see if they have converged to the same place. 

'''
import os, sys, shutil
from ase.io import read, write

from openpyxl import load_workbook

from tqdm import tqdm

adsorbate_name = sys.argv[1]
list_of_rows_from_argv = sys.argv[2:]
if len(list_of_rows_from_argv) == 0:
	exit()

list_of_rows = []
for index in range(len(list_of_rows_from_argv)):
	entry = list_of_rows_from_argv[index]
	if '-' in entry:
		start_no, end_no = entry.split('-')
		numbers = list(range(int(start_no),int(end_no)+1,1))
		list_of_rows += numbers
	else:
		list_of_rows.append(int(entry))
	
list_of_rows = sorted(list_of_rows)

System_Adsorbate_VASP_Jobs_folder_name = 'Part_C_Selected_Systems_with_Adsorbed_Species_to_Run_in_VASP'
if not System_Adsorbate_VASP_Jobs_folder_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(System_Adsorbate_VASP_Jobs_folder_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+System_Adsorbate_VASP_Jobs_folder_name+' folder is.')
    print('This program will finish without beginning.')
    exit()

Run_AdsorberPY_name = 'Run_Adsorber.py'
if not Run_AdsorberPY_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(Run_AdsorberPY_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+Run_AdsorberPY_name+' script is.')
    print('This program will finish without beginning.')
    exit()

save_folder_name = 'Part_D_Results_Folder'    
excel_name = "Part_D_Information_on_VASP_Calculations.xlsx"
workbook = load_workbook(save_folder_name+'/'+excel_name)

submission_folder_name = 'Submission_Folder'
similar_systems_name = 'Similar_Systems'
Part_D_Results_Folder_name = 'Part_D_Results_Folder'
saving_suffix_path = save_folder_name+'/'+similar_systems_name

sheetname = str(adsorbate_name)
# ------------------------------------------------------------------------------------------------
sheet = workbook[sheetname]
print(sheetname)
clusters_info = []
energies = []
clusters = []
for row_index in list_of_rows:
	job_name = sheet.cell(row=row_index, column=3).value
	job_path = sheet.cell(row=row_index, column=4).value
	energy = float(sheet.cell(row=row_index, column=12).value.replace('=',''))
	energies.append(energy)
	cluster = read(job_path+'/OUTCAR')
	clusters_info.append((job_name,energy))
	cluster.get_potential_energy()
	clusters.append(cluster)

min_energy = min(energies)
# ------------------------------------------------------------------------------------------------

CI_path = save_folder_name+'/Clusters_Info_'+str(adsorbate_name)
if os.path.exists(CI_path):
	shutil.rmtree(CI_path)
os.makedirs(CI_path)

with open(CI_path+'/Clusters_Info_'+str(adsorbate_name)+'.txt','w') as CI_TXT:
	counter = 1
	for name, energy in clusters_info:
		CI_TXT.write(str(counter)+' ('+str(list_of_rows[counter-1])+'): '+str(name)+' ('+str(energy)+' eV, '+str(energy-min_energy)+' eV)\n')
		counter += 1

write(CI_path+'/Clusters_Info_'+str(adsorbate_name)+'.traj',clusters)

# ------------------------------------------------------------------------------------------------