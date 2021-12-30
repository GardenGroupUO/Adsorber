#!/usr/bin/env python3
'''
Geoffrey Weal, Run_Adsorber_compare_sytems_with_same_binding.py, 14/09/2021

This program is designed to allow you to look at different VASP jobs with adsorbates in different positions that have converged to similar places.

This program is designed to allow you to see if they have converged to the same place. 

'''
import os, sys, re
from ase.io import read, write

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from tqdm import tqdm

# -----------------------------------------------
def set_write_similarity_traj_files(input_value):
	true_values = ['true','t']
	false_values = ['false','f']
	if input_value.lower() in true_values+false_values:
		if input_value.lower() in false_values:
			write_similarity_traj_files = False
		else:
			write_similarity_traj_files = True
	else:
		write_similarity_traj_files = False
	return write_similarity_traj_files

def is_float(element):
    try:
        float(element)
        return True
    except ValueError:
        return False

def set_upper_energy_limit(input_value):
	if input_value.isdigit() or is_float(input_value):
		upper_energy_limit = float(input_value)
	else:
		upper_energy_limit = float('inf')
		print('Upper Energy value not set (default).')
	return upper_energy_limit

if len(sys.argv) == 1:
	write_similarity_traj_files = set_write_similarity_traj_files(None)
	upper_energy_limit = set_upper_energy_limit(None)
elif len(sys.argv) == 2:
	write_similarity_traj_files = set_write_similarity_traj_files(sys.argv[1])
	upper_energy_limit = set_upper_energy_limit(sys.argv[1])
elif len(sys.argv) >= 3:
	write_similarity_traj_files = set_write_similarity_traj_files(sys.argv[1])
	upper_energy_limit = set_upper_energy_limit(sys.argv[2])
# -----------------------------------------------

def make_dir(save_folder_name):
	if not os.path.exists(save_folder_name):
		os.makedirs(save_folder_name)

System_Adsorbate_VASP_Jobs_folder_name = 'Part_C_Selected_Systems_with_Adsorbed_Species_to_Run_in_VASP'
if not System_Adsorbate_VASP_Jobs_folder_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(System_Adsorbate_VASP_Jobs_folder_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+System_Adsorbate_VASP_Jobs_folder_name+' folder is.')
    print('This program will finish without beginning.')
    exit()

Run_AdsorberPY_name = 'Run_Adsorber.py'
if not Run_AdsorberPY_name in os.listdir('.'):
    print('Error: Could not find the folder: '+str(Run_AdsorberPY_name))
    print('You need to execute Run_Adsorber_determine_unconverged_VASP_jobs.py in the same place as your '+Run_AdsorberPY_name+' script is.')
    print('This program will finish without beginning.')
    exit()

save_folder_name = 'Part_D_Results_Folder'    
excel_name = "Part_D_Information_on_VASP_Calculations.xlsx"
workbook = load_workbook(save_folder_name+'/'+excel_name)

sheetnames = workbook.sheetnames
if 'Originals' in sheetnames:
	sheetnames.remove('Originals')

def get_key_name(key):
	key_split = key.replace(' ','_')
	key_split = key_split.rstrip().split('(,')
	for index in range(len(key_split)):
		key_split_part = key_split[index].replace('[','')
		key_split_part = key_split_part.replace(']','')
		key_split_part = key_split_part.replace('(','')
		key_split_part = key_split_part.replace(')','')
		key_split_part = key_split_part.replace('->','to')
		key_split[index] = key_split_part.replace(',','+')
	key_name = '__'.join(key_split)
	return key_name

submission_folder_name = 'Submission_Folder'
similar_systems_name = 'Similar_Systems'
Part_D_Results_Folder_name = 'Part_D_Results_Folder'
saving_suffix_path = save_folder_name+'/'+similar_systems_name
for sheetname in sheetnames:
	# ------------------------------------------------------------------------------------------------
	sheet = workbook[sheetname]
	print(sheetname)
	system_information = []
	for row_index in range(2,sheet.max_row+1):
		job_name = sheet.cell(row=row_index, column=3).value
		job_path = sheet.cell(row=row_index, column=4).value
		energy = float(sheet.cell(row=row_index, column=12).value.replace('=',''))
		binding_surface_atoms = sheet.cell(row=row_index, column=16).value
		system_information.append([energy,binding_surface_atoms,job_name,job_path,row_index])
	system_information.sort()
	# ------------------------------------------------------------------------------------------------
	similar_systems = {}
	for energy, binding_surface_atoms, job_name, job_path, row_index in system_information:
		similar_systems.setdefault(binding_surface_atoms,[]).append((energy,row_index,job_name,job_path))
	# ------------------------------------------------------------------------------------------------
	energy_of_each_similar_type = []
	for key in similar_systems.keys():
		lowest_energy, row_index, name, path = min(similar_systems[key])
		energy_of_each_similar_type.append([lowest_energy,row_index,key,len(similar_systems[key]),path])
	energy_of_each_similar_type.sort()
	print('-----------------------------------------------------')
	print('Types of similar systems')
	print('row_index\tsimilar_system filename\t[number of ]\t(energy)')
	counter = 1
	lowest_lowest_energy = energy_of_each_similar_type[0][0]
	with open(Part_D_Results_Folder_name+'/'+similar_systems_name+'_'+sheetname+'.txt','w') as all_sim_systems_file_representative:
		for lowest_energy, row_index, key, no_of_sims, path in energy_of_each_similar_type:
			if lowest_energy <= (lowest_lowest_energy+upper_energy_limit):
				if key == None:
					print('Error: '+str(path)+' may not have the adsorbate binding to the surface of your system. To check. Will ignore this.')
					continue
				to_string = str(row_index)+': '+str(get_key_name(key))+'\t['+str(no_of_sims)+']\t'+str(round(lowest_energy-lowest_lowest_energy,3))+' eV ('+str(round(lowest_energy,3))+' eV)'
				print(to_string)
				all_sim_systems_file_representative.write(str(path)+' \t'+to_string+'\n')
				counter += 1
			else:
				break
	print('-----------------------------------------------------')
	# ------------------------------------------------------------------------------------------------
	if write_similarity_traj_files:
		pbar = tqdm(energy_of_each_similar_type,unit="Similar systems processed")
		for _, _, key, _, _ in pbar:
			if len(similar_systems[key]) > 1:
				similar_systems[key].sort()
				key_name = get_key_name(key)
				saving_path = saving_suffix_path+'/'+sheetname+'/'+key_name
				pbar.set_description("Processing %s" % key_name)
				#print(key_name)
				make_dir(saving_path)
				file = open(saving_path+'/similar_systems.txt','w')
				file.write('no.\tname\tenergy (eV)\trow in excel spreadsheet\n')
				all_outcar_files = []
				counter = 1
				for energy, row_index, job_name, job_path in similar_systems[key]:
					if os.path.exists(job_path+'/OUTCAR'):
						final_image = read(job_path+'/OUTCAR')
					elif os.path.exists(job_path+'/CONTCAR'):
						final_image = read(job_path+'/CONTCAR')
					else:
						submission_folders = [folder for folder in os.listdir(job_path) if (os.path.isdir(job_path+'/'+folder) and (submission_folder_name in folder) and (not 'Issue' in folder))]
						if len(submission_folders) > 0:
							latest_submission_folder = max(submission_folders,key=lambda foldername: int(foldername.replace(submission_folder_name+'_','')))
							if os.path.exists(job_path+'/'+latest_submission_folder+'/OUTCAR'):
								final_image = read(job_path+'/'+latest_submission_folder+'/OUTCAR')
							elif os.path.exists(job_path+'/'+latest_submission_folder+'/CONTCAR'):
								final_image = read(job_path+'/'+latest_submission_folder+'/CONTCAR')
							else:
								error_message(job_path)
						else:
							error_message(job_path)
					#write(saving_path+'/'+job_name+'.xyz',final_image)
					all_outcar_files.append(final_image)
					file.write(str(counter)+'\t'+job_name+'\t'+str(round(float(energy),5))+'\t'+str(row_index)+'\n')
					counter += 1
				write(saving_path+'/'+key_name+'.traj',all_outcar_files)
	# ------------------------------------------------------------------------------------------------

def error_message(job_path):
	print('Error in Run_Adsorber_compare_systems_with_same_binding.py')
	print('Can not find a OUTCAR or CONTCAR in this folder or any Submission_Folders that did not have issues')
	print(job_path)
	print('Check this out')
	exit('This program will not exit without completing.')